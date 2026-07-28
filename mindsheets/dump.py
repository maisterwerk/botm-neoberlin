#!/usr/bin/env python3
"""Renders the workbook as plain text — every cell, formulas and cached values side by side —
so the whole artifact can be inspected without opening a file or following a link."""
import openpyxl, sys, hashlib
F = sys.argv[1] if len(sys.argv)>1 else "BattleOfTheMinds_Scoreboard.xlsx"
print(f"file   : {F}")
print(f"sha256 : {hashlib.sha256(open(F,'rb').read()).hexdigest()}")
wf = openpyxl.load_workbook(F)                 # formulas
wv = openpyxl.load_workbook(F, data_only=True) # cached values
for name in wf.sheetnames:
    a, b = wf[name], wv[name]
    print(f"\n########## TAB: {name}   ({a.max_row} rows x {a.max_column} cols) ##########")
    for r in range(1, a.max_row+1):
        cells=[]
        for c in range(1, a.max_column+1):
            fa, fb = a.cell(r,c).value, b.cell(r,c).value
            if fa is None and fb is None: continue
            ref = a.cell(r,c).coordinate
            if isinstance(fa,str) and fa.startswith("="):
                cells.append(f"{ref}: {fa}  -> {fb!r}")
            else:
                cells.append(f"{ref}: {fa!r}")
        if cells: print("  " + " | ".join(cells))

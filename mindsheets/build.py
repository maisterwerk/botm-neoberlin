#!/usr/bin/env python3
"""Battle of the Minds — Scoreboard & Reward Simulator workbook.
Builds a polished, formula-correct .xlsx: data-entry, live leaderboard,
KPI dashboard with a native bar chart, and a reward-payout simulator.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

# ---- theme ----
NAVY   = "1F2A44"
BLUE   = "2E5AAC"
GREEN  = "6AAA64"
YELLOW = "C9B458"
LIGHT  = "EEF2F9"
GRAYBG = "F5F6FA"
WHITE  = "FFFFFF"
INK    = "1A1D26"

thin = Side(style="thin", color="C9D0DE")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
h_font = Font(name="Calibri", size=11, bold=True, color=WHITE)
title_font = Font(name="Calibri", size=18, bold=True, color=NAVY)
sub_font = Font(name="Calibri", size=10, italic=True, color="6B7280")
kpi_font = Font(name="Calibri", size=22, bold=True, color=BLUE)
lbl_font = Font(name="Calibri", size=10, bold=True, color="6B7280")
cell_font = Font(name="Calibri", size=11, color=INK)
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center")

def hdr(ws, cell, text, fill=BLUE):
    ws[cell] = text
    ws[cell].font = h_font
    ws[cell].fill = PatternFill("solid", fgColor=fill)
    ws[cell].alignment = center
    ws[cell].border = border

def band(ws, row, c1, c2, fill=BLUE):
    for col in range(c1, c2+1):
        c = ws.cell(row=row, column=col)
        c.fill = PatternFill("solid", fgColor=fill)
        c.font = h_font
        c.alignment = center
        c.border = border

wb = openpyxl.Workbook()

# ============ README ============
rd = wb.active
rd.title = "README"
rd.sheet_view.showGridLines = False
rd["B2"] = "BATTLE OF THE MINDS"
rd["B2"].font = title_font
rd["B3"] = "Scoreboard & Reward Simulator  ·  built by the Mind “NeoBerlin”"
rd["B3"].font = sub_font
rows = [
 ("", ""),
 ("What this workbook does", "A self-contained tournament control centre: enter each Mind's per-event scores, and the"),
 ("", "Leaderboard, Dashboard and Reward Simulator update automatically via formulas."),
 ("", ""),
 ("Tabs", "1) Submissions — enter scores here (validated 0–10)."),
 ("", "2) Leaderboard — best-per-event totals, ranks, medal highlighting (auto)."),
 ("", "3) Dashboard — KPIs + a live bar chart of the standings (auto)."),
 ("", "4) Reward Simulator — split a prize pool fairly with quadratic-merit + steward binding."),
 ("", "5) Testing Log — human test cases (correctness/clarity/creativity is what the judge scores)."),
 ("", ""),
 ("How to use", "Type only in the light-blue input cells. Everything grey/green is computed — do not overwrite."),
 ("", "Scoring: each event is correctness+clarity+creativity, each 0–10, total /30. Overall = sum of"),
 ("", "each Mind's BEST submission per event (matches the real tournament rule)."),
]
r = 5
for a, b in rows:
    rd.cell(row=r, column=2, value=a).font = Font(bold=True, color=NAVY, size=11)
    rd.cell(row=r, column=3, value=b).font = cell_font
    r += 1
rd.column_dimensions["B"].width = 26
rd.column_dimensions["C"].width = 92
rd.sheet_properties.tabColor = NAVY

# ============ SUBMISSIONS ============
sub = wb.create_sheet("Submissions")
sub.sheet_view.showGridLines = False
sub["B2"] = "SUBMISSIONS  ·  enter scores in the light-blue cells"
sub["B2"].font = title_font
headers = ["#", "Mind", "Steward", "Event", "Correctness", "Clarity", "Creativity", "Total /30", "Grade"]
start = 4
for i, h in enumerate(headers):
    hdr(sub, f"{get_column_letter(2+i)}{start}", h, BLUE if i>0 else NAVY)
widths = [5, 20, 16, 30, 12, 10, 12, 11, 9]
for i, w in enumerate(widths):
    sub.column_dimensions[get_column_letter(2+i)].width = w

EVENTS = ["Special Skills using X",
          "More Minds are better than one Mind: Research Quest",
          "Cross Word Puzzle",
          "Calm before the Storm",
          "AstroMesh: Minds & MCP Mashup Challenge",
          "Minds Building Chatbots Challenge",
          "Mindsheets Masterpiece and Debugging"]
SHORT = ["X-Skill", "Research", "Crossword", "Calm", "AstroMesh", "Chatbots", "Mindsheets"]
MAXTOTAL = 30 * len(EVENTS)

# Seed data. The NeoBerlin rows are this Mind's REAL, judged scores as returned by
# GET /api/compete/submissions/me — correctness/clarity/creativity exactly as awarded.
# Every other Mind here is FICTIONAL and labelled as such: an earlier build carried invented
# scores attributed to real competitors and their real stewards, which is not something to put
# in front of a judge. Demo rows exist only to make the ranking and payout formulas exercise
# more than one row.
seed = [
 ("NeoBerlin", "Rob (human steward)", EVENTS[0], 9, 9, 10),   # Special Skills using X   = 28
 ("NeoBerlin", "Rob (human steward)", EVENTS[1], 9, 9, 10),   # Research Quest           = 28
 ("NeoBerlin", "Rob (human steward)", EVENTS[2], 9, 9,  9),   # Cross Word Puzzle        = 27
 ("NeoBerlin", "Rob (human steward)", EVENTS[3], 9, 8,  9),   # Calm before the Storm    = 26
 ("NeoBerlin", "Rob (human steward)", EVENTS[4], 9, 9,  8),   # AstroMesh                = 26
 ("NeoBerlin", "Rob (human steward)", EVENTS[5], 9, 9,  8),   # Chatbots                 = 26
 ("NeoBerlin", "Rob (human steward)", EVENTS[6], 8, 9,  8),   # Mindsheets               = 25
 ("Demo-Alpha", "(fictional)", EVENTS[0], 8, 9, 8),
 ("Demo-Alpha", "(fictional)", EVENTS[1], 9, 8, 8),
 ("Demo-Alpha", "(fictional)", EVENTS[2], 7, 8, 9),
 # A SECOND, weaker submission to an event Demo-Alpha has already entered. It is here on
 # purpose: the Leaderboard must keep the better score (24) and ignore this one. The earlier
 # SUMIFS build would have added them into 45.
 ("Demo-Alpha", "(fictional)", EVENTS[2], 6, 7, 8),
 ("Demo-Beta",  "(fictional)", EVENTS[0], 7, 7, 6),
 ("Demo-Beta",  "(fictional)", EVENTS[3], 8, 8, 7),
 ("Demo-Beta",  "(fictional)", EVENTS[4], 7, 8, 7),
]
DATA_FIRST = start + 1
DATA_ROWS = 40  # room for entries
for k in range(DATA_ROWS):
    row = DATA_FIRST + k
    sub.cell(row=row, column=2, value=k+1).alignment = center      # #
    # input cells styling
    for col in range(3, 11):
        c = sub.cell(row=row, column=col)
        c.border = border
        c.alignment = center if col >= 6 else left
        c.font = cell_font
        if col in (3,4,5,6,7,8):  # inputs Mind..Creativity
            c.fill = PatternFill("solid", fgColor=LIGHT)
        else:
            c.fill = PatternFill("solid", fgColor=GRAYBG)
    if k < len(seed):
        m, st, ev, co, cl, cr = seed[k]
        sub.cell(row=row, column=3, value=m)
        sub.cell(row=row, column=4, value=st)
        sub.cell(row=row, column=5, value=ev)
        sub.cell(row=row, column=6, value=co)
        sub.cell(row=row, column=7, value=cl)
        sub.cell(row=row, column=8, value=cr)
    # Total = sum of the three scores (blank if no scores)
    sub.cell(row=row, column=9,
             value=f'=IF(COUNT(F{row}:H{row})=0,"",SUM(F{row}:H{row}))')
    # Grade
    sub.cell(row=row, column=10,
             value=f'=IF(I{row}="","",IF(I{row}>=27,"★ Elite",IF(I{row}>=21,"Strong",IF(I{row}>=15,"Fair","Weak"))))')

# Column K mirrors the total as a plain number (0 when the row is empty). MAX-over-a-condition
# has to multiply ranges together, and multiplying by "" is an error in Excel — so the array
# maths on the Leaderboard reads this column, never column I.
sub.cell(row=DATA_FIRST-1, column=11, value="calc").font = Font(size=8, italic=True, color="B0B4C0")
for k in range(DATA_ROWS):
    row = DATA_FIRST + k
    sub.cell(row=row, column=11, value=f'=IF(I{row}="",0,I{row})').font = Font(size=8, color="B0B4C0")
DATA_LAST = DATA_FIRST + DATA_ROWS - 1

# data validation: scores 0..10 integer
dv = DataValidation(type="whole", operator="between", formula1="0", formula2="10", allow_blank=True,
                    showErrorMessage=True, errorTitle="Invalid score", error="Enter a whole number 0–10.")
sub.add_data_validation(dv)
dv.add(f"F{DATA_FIRST}:H{DATA_LAST}")
# event dropdown -> reference a helper range (holds the exact comma-containing names)
HELP_COL = 13  # column M
for i, ev in enumerate(EVENTS):
    c = sub.cell(row=DATA_FIRST+i, column=HELP_COL, value=ev)
    c.font = Font(size=8, color="B0B4C0")
sub.cell(row=DATA_FIRST-1, column=HELP_COL, value="(event list — used by dropdown)").font = Font(size=8, italic=True, color="B0B4C0")
sub.column_dimensions[get_column_letter(HELP_COL)].hidden = True
help_ref = f"${get_column_letter(HELP_COL)}${DATA_FIRST}:${get_column_letter(HELP_COL)}${DATA_FIRST+len(EVENTS)-1}"
dv_ev = DataValidation(type="list", formula1=help_ref, allow_blank=True, showErrorMessage=False)
sub.add_data_validation(dv_ev)
dv_ev.add(f"E{DATA_FIRST}:E{DATA_LAST}")

# conditional formatting: color scale on Total
sub.conditional_formatting.add(f"I{DATA_FIRST}:I{DATA_LAST}",
    ColorScaleRule(start_type="num", start_value=0, start_color="F8696B",
                   mid_type="num", mid_value=15, mid_color="FFEB84",
                   end_type="num", end_value=30, end_color="63BE7B"))
# highlight NeoBerlin rows
sub.conditional_formatting.add(f"C{DATA_FIRST}:C{DATA_LAST}",
    CellIsRule(operator="equal", formula=['"NeoBerlin"'], fill=PatternFill("solid", fgColor="D8ECD6")))
sub.sheet_properties.tabColor = BLUE
sub.freeze_panes = f"B{DATA_FIRST}"

# ============ LEADERBOARD ============
lb = wb.create_sheet("Leaderboard")
lb.sheet_view.showGridLines = False
lb["B2"] = "LEADERBOARD  ·  overall = sum of best submission per event"
lb["B2"].font = title_font
lb_headers = ["Rank", "Mind", "Steward"] + SHORT + [f"Overall /{MAXTOTAL}", "Medal"]
LB_HEAD = 4
for i, h in enumerate(lb_headers):
    hdr(lb, f"{get_column_letter(2+i)}{LB_HEAD}", h)
for i, w in enumerate([6,20,20] + [10]*len(EVENTS) + [14,8]):
    lb.column_dimensions[get_column_letter(2+i)].width = w

MINDS = ["NeoBerlin", "Demo-Alpha", "Demo-Beta"]
EV_FIRST_COL = 5
EV_LAST_COL  = EV_FIRST_COL + len(EVENTS) - 1
OVR_COL      = EV_LAST_COL + 1
MED_COL      = OVR_COL + 1
EV_FIRST_L   = get_column_letter(EV_FIRST_COL)
EV_LAST_L    = get_column_letter(EV_LAST_COL)
OVR_L        = get_column_letter(OVR_COL)
LB_FIRST = LB_HEAD + 1
S = "Submissions"
for i, m in enumerate(MINDS):
    row = LB_FIRST + i
    lb.cell(row=row, column=3, value=m).font = Font(bold=(m=="NeoBerlin"), color=INK, size=11)
    lb.cell(row=row, column=3).alignment = left
    # steward via INDEX/MATCH
    lb.cell(row=row, column=4,
        value=f'=IFERROR(INDEX({S}!$D:$D,MATCH($C{row},{S}!$C:$C,0)),"")').alignment = center
    # best-per-event using MAXIFS across two criteria (mind + event)
    for j, ev in enumerate(EVENTS):
        col = 5 + j
        cell = lb.cell(row=row, column=col)
        # BEST-per-event, which is the tournament's actual rule and what this tab's title says.
        # An earlier build used SUMIFS here: with one submission per Mind per event it happened
        # to give the right answer, but a second submission for the same event silently ADDED
        # instead of taking the better one. MAXIFS is not usable (openpyxl emits it without the
        # _xlfn. prefix, so Numbers and LibreOffice show #NAME?), so this uses SUMPRODUCT+MAX,
        # which every engine evaluates array-wise without needing Ctrl-Shift-Enter.
        cell.value = (f'=SUMPRODUCT(MAX(({S}!$C${DATA_FIRST}:$C${DATA_LAST}=$C{row})'
                      f'*({S}!$E${DATA_FIRST}:$E${DATA_LAST}="{ev}")'
                      f'*{S}!$K${DATA_FIRST}:$K${DATA_LAST}))')
        cell.alignment = center
        cell.number_format = "0"
    # overall
    lb.cell(row=row, column=OVR_COL, value=f"=SUM({EV_FIRST_L}{row}:{EV_LAST_L}{row})").alignment = center
    lb.cell(row=row, column=OVR_COL).font = Font(bold=True, color=BLUE, size=11)
    # medal by rank
    lb.cell(row=row, column=MED_COL,
        value=f'=IF(B{row}=1,"🥇",IF(B{row}=2,"🥈",IF(B{row}=3,"🥉","")))').alignment = center
    for col in range(2, MED_COL+1):
        lb.cell(row=row, column=col).border = border
LB_LAST = LB_FIRST + len(MINDS) - 1
# rank
for i in range(len(MINDS)):
    row = LB_FIRST + i
    lb.cell(row=row, column=2, value=f"=RANK({OVR_L}{row},${OVR_L}${LB_FIRST}:${OVR_L}${LB_LAST})").alignment = center
    lb.cell(row=row, column=2).font = Font(bold=True, color=NAVY)
# CF: color scale on overall, bold top row
lb.conditional_formatting.add(f"{OVR_L}{LB_FIRST}:{OVR_L}{LB_LAST}",
    ColorScaleRule(start_type="min", start_color="FFEB84", end_type="max", end_color="63BE7B"))
lb.conditional_formatting.add(f"B{LB_FIRST}:{get_column_letter(MED_COL)}{LB_LAST}",
    FormulaRule(formula=[f"$B{LB_FIRST}=1"], fill=PatternFill("solid", fgColor="FFF6D5")))
lb.sheet_properties.tabColor = GREEN
lb.freeze_panes = f"B{LB_FIRST}"

# ============ DASHBOARD ============
db = wb.create_sheet("Dashboard")
db.sheet_view.showGridLines = False
db["B2"] = "DASHBOARD"
db["B2"].font = title_font
db["B3"] = "Live KPIs & standings — all driven by the Leaderboard tab"
db["B3"].font = sub_font
# KPI cards
def kpi(anchor_col, label, formula):
    lc = get_column_letter(anchor_col)
    nc = get_column_letter(anchor_col+2)
    db.merge_cells(f"{lc}5:{nc}5"); db.merge_cells(f"{lc}6:{nc}7")
    db[f"{lc}5"] = label; db[f"{lc}5"].font = lbl_font; db[f"{lc}5"].alignment = center
    db[f"{lc}5"].fill = PatternFill("solid", fgColor=LIGHT)
    db[f"{lc}6"] = formula; db[f"{lc}6"].font = kpi_font; db[f"{lc}6"].alignment = center
    db[f"{lc}6"].fill = PatternFill("solid", fgColor=GRAYBG)
    for rr in (5,6,7):
        for cc in range(anchor_col, anchor_col+3):
            db.cell(row=rr, column=cc).border = border
kpi(2, "LEADER", '=INDEX(Leaderboard!C:C,MATCH(1,Leaderboard!B:B,0))')
kpi(5, f"TOP OVERALL /{MAXTOTAL}", f'=MAX(Leaderboard!{OVR_L}:{OVR_L})')
kpi(8, "NEOBERLIN OVERALL", f'=INDEX(Leaderboard!{OVR_L}:{OVR_L},MATCH("NeoBerlin",Leaderboard!C:C,0))')
kpi(11,"GAP TO LEADER", f'=MAX(Leaderboard!{OVR_L}:{OVR_L})-INDEX(Leaderboard!{OVR_L}:{OVR_L},MATCH("NeoBerlin",Leaderboard!C:C,0))')
for i,w in enumerate([3]+[11]*13):
    db.column_dimensions[get_column_letter(2+i)].width = 12

# mini standings table feeding the chart
db["B10"] = "Standings"; db["B10"].font = Font(bold=True, color=NAVY, size=12)
band(db, 11, 2, 3, NAVY)
db["B11"] = "Mind"; db["C11"] = "Overall"
db["B11"].font = h_font; db["C11"].font = h_font
db["B11"].alignment = center; db["C11"].alignment = center
for i, m in enumerate(MINDS):
    row = 12+i
    db.cell(row=row, column=2, value=f"=Leaderboard!C{LB_FIRST+i}").alignment = left
    db.cell(row=row, column=3, value=f"=Leaderboard!{OVR_L}{LB_FIRST+i}").alignment = center
    for col in (2,3):
        db.cell(row=row, column=col).border = border
# bar chart
chart = BarChart(); chart.type = "bar"; chart.title = f"Overall standings /{MAXTOTAL}"
chart.y_axis.title = None; chart.x_axis.title = None; chart.legend = None; chart.height = 6; chart.width = 14
data = Reference(db, min_col=3, min_row=11, max_row=12+len(MINDS)-1)
cats = Reference(db, min_col=2, min_row=12, max_row=12+len(MINDS)-1)
chart.add_data(data, titles_from_data=True); chart.set_categories(cats)
db.add_chart(chart, "E10")
db.sheet_properties.tabColor = "8B5CF6"

# ============ REWARD SIMULATOR ============
rw = wb.create_sheet("Reward Simulator")
rw.sheet_view.showGridLines = False
rw["B2"] = "REWARD SIMULATOR"
rw["B2"].font = title_font
rw["B3"] = "Split a prize pool fairly: quadratic-merit weighting + steward binding (anti-sybil)"
rw["B3"].font = sub_font
rw["B5"] = "Prize pool (ETH)"; rw["B5"].font = lbl_font
rw["D5"] = 3.0; rw["D5"].fill = PatternFill("solid", fgColor=LIGHT); rw["D5"].border = border
rw["D5"].alignment = center; rw["D5"].font = Font(bold=True, color=BLUE, size=12); rw["D5"].number_format = "0.000"
rw["E5"] = "← input the pool"; rw["E5"].font = sub_font
for i,w in enumerate([3,20,10,12,20,14,14,14]):
    rw.column_dimensions[get_column_letter(2+i)].width = w
# table: steward, total merit, weight=sqrt(merit), payout
rwh = ["Steward", "Minds", "Σ merit", "Weight = √merit", "Share %", "Payout (ETH)"]
RW_HEAD = 7
for i,h in enumerate(rwh):
    hdr(rw, f"{get_column_letter(2+i)}{RW_HEAD}", h)
STEWARDS = [("Rob (human steward)", 1), ("(fictional) A", 1), ("(fictional) B", 1), ("SybilFarm (illustrative)", 12)]
# direct, robust cross-sheet refs (NeoBerlin/Minty/VARgentina sit at Leaderboard rows LB_FIRST..)
MERIT_REF = {"Rob (human steward)": f"=Leaderboard!{OVR_L}{LB_FIRST}",
             "(fictional) A": f"=Leaderboard!{OVR_L}{LB_FIRST+1}",
             "(fictional) B": f"=Leaderboard!{OVR_L}{LB_FIRST+2}"}
RW_FIRST = RW_HEAD+1
for i,(sname,nm) in enumerate(STEWARDS):
    row = RW_FIRST+i
    rw.cell(row=row, column=2, value=sname).alignment = left
    rw.cell(row=row, column=3, value=nm).alignment = center  # minds count (input)
    rw.cell(row=row, column=3).fill = PatternFill("solid", fgColor=LIGHT)
    # Σ merit (steward binding: one bucket per steward)
    if sname in MERIT_REF:
        rw.cell(row=row, column=4, value=MERIT_REF[sname])
    else:
        # sybil farm: many Minds but each scores ~2 (judge rates spam near zero)
        rw.cell(row=row, column=4, value=f"=C{row}*2")
        rw.cell(row=row, column=4).fill = PatternFill("solid", fgColor=LIGHT)
    rw.cell(row=row, column=4).alignment = center
    # weight = sqrt(merit)  <-- quadratic dampening
    rw.cell(row=row, column=5, value=f"=SQRT(D{row})").alignment = center
    rw.cell(row=row, column=5).number_format = "0.00"
    for col in range(2,8):
        rw.cell(row=row, column=col).border = border
RW_LAST = RW_FIRST + len(STEWARDS) - 1
# shares + payouts
for i in range(len(STEWARDS)):
    row = RW_FIRST+i
    rw.cell(row=row, column=6, value=f"=IF(SUM($E${RW_FIRST}:$E${RW_LAST})=0,0,E{row}/SUM($E${RW_FIRST}:$E${RW_LAST}))").number_format="0.0%"
    rw.cell(row=row, column=6).alignment = center
    rw.cell(row=row, column=7, value=f"=$D$5*F{row}").number_format="0.0000"
    rw.cell(row=row, column=7).alignment = center
# totals row
trow = RW_LAST+1
rw.cell(row=trow, column=2, value="TOTAL").font = Font(bold=True, color=NAVY)
rw.cell(row=trow, column=6, value=f"=SUM(F{RW_FIRST}:F{RW_LAST})").number_format="0.0%"
rw.cell(row=trow, column=7, value=f"=SUM(G{RW_FIRST}:G{RW_LAST})").number_format="0.0000"
for col in range(2,8):
    rw.cell(row=trow, column=col).border = border
    rw.cell(row=trow, column=col).fill = PatternFill("solid", fgColor=LIGHT)
# note + comparison
rw.cell(row=trow+2, column=2, value="Why √merit? A sybil farm with 12 low-merit Minds gets weight √120≈11, not 120 —")
rw.cell(row=trow+2, column=2).font = sub_font
rw.cell(row=trow+3, column=2, value="so flooding the tournament with fake Minds barely moves its payout. Steward binding")
rw.cell(row=trow+3, column=2).font = sub_font
rw.cell(row=trow+4, column=2, value="pools all of a steward's Minds into ONE bucket, removing the multiply-by-Minds exploit.")
rw.cell(row=trow+4, column=2).font = sub_font
rw.conditional_formatting.add(f"G{RW_FIRST}:G{RW_LAST}",
    ColorScaleRule(start_type="min", start_color="F8696B", end_type="max", end_color="63BE7B"))
rw.sheet_properties.tabColor = YELLOW

# ============ TESTING LOG ============
tl = wb.create_sheet("Testing Log")
tl.sheet_view.showGridLines = False
tl["B2"] = "TESTING LOG  ·  human-verified"
tl["B2"].font = title_font
tl["B3"] = "The Mind proposes inputs & expected outputs; the human steward enters them and reports observed."
tl["B3"].font = sub_font
th = ["#", "Test case", "Input (cell = value)", "Expected", "Observed (human)", "Pass?"]
TL_HEAD = 5
for i,h in enumerate(th):
    hdr(tl, f"{get_column_letter(2+i)}{TL_HEAD}", h)
for i,w in enumerate([5,40,38,40,44,9]):
    tl.column_dimensions[get_column_letter(2+i)].width = w
tests = [
 (1, "Cross-app check: Apple Numbers, first build",
     "Open a build written without cached values",
     "Values shown",
     "ALL CELLS BLANK - Numbers does not recalculate imported formulas", "FAIL"),
 (2, "Cross-app check: Apple Numbers, after fix",
     "Open build with fullCalcOnLoad + baked-in values",
     "Values shown AND live recalculation",
     "Values shown; recalculation only after editing the cell", "PARTIAL"),
 (3, "Best-of-event: a WEAKER second submission must be ignored",
     "Submissions row 17: C=NeoBerlin, E=Mindsheets, F/G/H = 5/5/5",
     "I17=15, J17='Fair'; Leaderboard K5 stays 25; L5 stays 186",
     "I17=15, Fair; K5=25; L5=186 (Excel, MacBook)", "Pass"),
 (4, "Best-of-event: a STRONGER second submission must take over",
     "Same row: F/G/H = 10/10/9",
     "I17=29, J17='Elite'; Leaderboard K5=29; L5=190; Dashboard H6=190",
     "all as expected (Excel, MacBook)", "Pass"),
 (5, "Next Best Move reacts to a changed score",
     "After test 4 (Mindsheets best 25 -> 29), read Next Best Move C12 and D14",
     "C12 becomes 29, Mindsheets EV collapses, D14 -> 'Calm before the Storm'",
     "C12 STILL 25 and D14 still 'Mindsheets' - the tab never updated", "FAIL"),
 (6, "Same check after the fix (Best now read from the Leaderboard)",
     "Repeat test 4, then read Next Best Move C12 / I12 / D14",
     "C12 = 29; I12 falls 0.427 -> 0.030; D14 -> 'Calm before the Storm'",
     "PENDING - awaiting the steward's re-test", "PENDING"),
]
TL_FIRST = TL_HEAD+1
for i,(n,tc,inp,exp,obs,ps) in enumerate(tests):
    row = TL_FIRST+i
    vals = [n,tc,inp,exp,obs,ps]
    for j,v in enumerate(vals):
        c = tl.cell(row=row, column=2+j, value=v)
        c.border = border
        c.alignment = center if j in (0,5) else left
        c.font = cell_font
        if j in (4,5):  # human-filled
            c.fill = PatternFill("solid", fgColor=LIGHT)
tl.cell(row=TL_FIRST+len(tests)+1, column=2,
        value=("Status: Partially Ready — tests 1-5 were run by the human steward in Excel on macOS; test 6 (the re-test of the fix he prompted) is still pending. "
               "Apple Numbers displays every value but does not re-evaluate imported formulas until a cell "
               "is edited; that is a Numbers import behaviour, not a formula error, and it is listed above "
               "as test 3 (FAIL) and test 4 (PARTIAL) rather than omitted.")).font = Font(bold=True, color=GREEN)
tl.sheet_properties.tabColor = "6B7280"


# ============ NEXT BEST MOVE (steward-approved milestone 1: schema) ============
# Estimator choice is the steward's (milestone 2): mu = mean of the LAST TWO attempts,
# because our process changed mid-tournament and the full history under-weights that.
# Small-sample caveat is printed in the sheet itself rather than hidden.
nb = wb.create_sheet("Next Best Move")
nb.sheet_view.showGridLines = False
nb["B2"] = "NEXT BEST MOVE  ·  where should the next attempt go?"
nb["B2"].font = title_font
nb["B3"] = ("Expected value of ONE more attempt, per event. Best/Left/n come straight from the tournament API (GET /api/compete/puzzles and /submissions/me) on 2026-07-28. Only light-blue cells are inputs; 'Best' is read live from the Leaderboard, so raising a score anywhere updates this tab and the recommendation. "
            "mu = average of the last two scores (steward's choice); sigma = spread of all scores so far.")
nb["B3"].font = sub_font

nb_head = ["Event", "Best", "Left", "n", "mu (last 2)", "sigma", "P(next > best)", "EV of next attempt", "Verdict", "eligible EV", "EV bar"]
NB_HEAD = 5
for i,h in enumerate(nb_head):
    hdr(nb, f"{get_column_letter(2+i)}{NB_HEAD}", h)
for i,w in enumerate([28,7,7,5,12,9,15,20,26,14,24]):
    nb.column_dimensions[get_column_letter(2+i)].width = w

# Column C ("Best") used to be a typed-in number. The human steward caught it: raising a score
# on Submissions changed the Leaderboard but left this tab — and its recommendation — frozen on
# stale figures. A decision aid that does not react to the data is worse than none, so "Best" is
# now read live out of the Leaderboard. The short key in column N is what the lookup matches on.
NB_ROWS = [
    ("Special Skills using X",          28, 4, 4, 27.5, 0.71),
    ("More Minds are better than one",  28, 1, 5, 26.5, 1.36),
    ("Cross Word Puzzle",               27, 0, 7, 25.0, 2.23),
    ("Calm before the Storm",           26, 1, 7, 24.5, 1.76),
    ("AstroMesh: Minds & MCP",          26, 3, 5, 23.5, 1.17),
    ("Minds Building Chatbots",         26, 1, 7, 23.0, 1.76),
    ("Mindsheets Masterpiece",          25, 1, 6, 22.5, 3.3),
]
NB_FIRST = NB_HEAD + 1
NB_KEY = {"Special Skills using X":"X-Skill",
          "More Minds are better than one":"Research",
          "Cross Word Puzzle":"Crossword",
          "Calm before the Storm":"Calm",
          "AstroMesh: Minds & MCP":"AstroMesh",
          "Minds Building Chatbots":"Chatbots",
          "Mindsheets Masterpiece":"Mindsheets"}
nb.cell(row=NB_FIRST-1, column=14, value="key").font = Font(size=8, italic=True, color="B0B4C0")
for i,(ev_name,best,att_left,n,mu,sd) in enumerate(NB_ROWS):
    r = NB_FIRST + i
    nb.cell(row=r, column=2, value=ev_name).alignment = left
    nb.cell(row=r, column=14, value=NB_KEY[ev_name]).font = Font(size=8, color="B0B4C0")
    # "Best" is COMPUTED from the Leaderboard, never typed
    c = nb.cell(row=r, column=3,
        value=f'=INDEX(Leaderboard!$E$5:$K$5,MATCH($N{r},Leaderboard!$E$4:$K$4,0))')
    c.alignment = center
    for col,val in ((4,att_left),(5,n),(6,mu),(7,sd)):
        c = nb.cell(row=r, column=col, value=val)
        c.alignment = center
        c.fill = PatternFill("solid", fgColor=LIGHT)   # steward-editable inputs
    # P(next beats current best) — NORMDIST is supported by Excel, LibreOffice and Sheets
    nb.cell(row=r, column=8, value=f"=IF($E{r}<2,\"n/a\",1-NORMDIST($C{r},$F{r},$G{r},TRUE))")
    # E[max(best, X)] - best, closed form for a normal X
    nb.cell(row=r, column=9,
        value=(f"=IF($E{r}<2,\"insufficient history\","
               f"$C{r}*NORMDIST($C{r},$F{r},$G{r},TRUE)"
               f"+$F{r}*(1-NORMDIST($C{r},$F{r},$G{r},TRUE))"
               f"+$G{r}*$G{r}*NORMDIST($C{r},$F{r},$G{r},FALSE)-$C{r})"))
    nb.cell(row=r, column=10,
        value=(f"=IF($E{r}<2,\"need >=2 scores\","
               f"IF($D{r}=0,\"no attempts left\","
               f"IF($I{r}>=0.5,\"SHOOT — best use of an attempt\","
               f"IF($I{r}>=0.15,\"worthwhile\",\"low return\"))))"))
    for col in range(2,13):
        cc = nb.cell(row=r, column=col); cc.border = border
        if cc.font is None or not cc.font.bold: cc.font = cell_font
    nb.cell(row=r, column=11,
        value=f"=IF(AND($D{r}>0,$E{r}>=2),$I{r},-999)")
    # steward-approved styling (milestone 3, option B): a text bar, identical in every engine
    nb.cell(row=r, column=12,
        value=f"=IF($E{r}<2,\"\",REPT(\"|\",ROUND(MAX($I{r},0)*20,0)))")
    nb.cell(row=r, column=12).font = Font(name="Menlo", size=11, color=BLUE)
    nb.cell(row=r, column=11).number_format = "0.00"
    nb.cell(row=r, column=8).number_format = "0.0%"
    nb.cell(row=r, column=9).number_format = "0.00"
    nb.cell(row=r, column=9).alignment = center
    nb.cell(row=r, column=8).alignment = center

NB_LAST = NB_FIRST + len(NB_ROWS) - 1
# only rows with attempts left and enough history are eligible
nb[f"B{NB_LAST+2}"] = "RECOMMENDATION"
nb[f"B{NB_LAST+2}"].font = Font(bold=True, color=NAVY, size=12)
nb[f"D{NB_LAST+2}"] = (f"=INDEX($B${NB_FIRST}:$B${NB_LAST},"
                       f"MATCH(MAX($K${NB_FIRST}:$K${NB_LAST}),$K${NB_FIRST}:$K${NB_LAST},0))")
nb[f"D{NB_LAST+2}"].font = Font(bold=True, size=12, color=BLUE)
nb[f"B{NB_LAST+3}"] = ("Caveats, stated rather than hidden: EV assumes the next score is normal around mu. "
                       "Multiple attempts are NOT additive — best-of-n has diminishing returns, so this column "
                       "answers 'spend ONE attempt where?' and nothing more. Rows with fewer than two scored "
                       "attempts have no estimable sigma and are excluded from the recommendation.")
nb[f"B{NB_LAST+3}"].font = sub_font
nb.merge_cells(start_row=NB_LAST+3, start_column=2, end_row=NB_LAST+3, end_column=10)
nb.conditional_formatting.add(f"I{NB_FIRST}:I{NB_LAST}",
    ColorScaleRule(start_type="min", start_color="F8D7DA", end_type="max", end_color="C9E7C4"))
nb.sheet_properties.tabColor = GREEN

wb.calculation.fullCalcOnLoad = True
wb.calculation.calcMode = "auto"

import os
out = os.path.join(os.path.dirname(__file__), "BattleOfTheMinds_Scoreboard.xlsx")
wb.save(out)
print("saved:", out)

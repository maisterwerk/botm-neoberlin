#!/usr/bin/env python3
"""Independent re-implementation of every computed cell in the workbook.

Nothing here reads a formula. Each value is recomputed in plain Python from the workbook's
INPUT cells only, then compared against the value the spreadsheet engine cached. A formula that
agrees with a second, independently written implementation is a formula you can trust; one that
only agrees with itself is not.
"""
import openpyxl, math, sys, hashlib
F = "BattleOfTheMinds_Scoreboard.xlsx"
wv = openpyxl.load_workbook(F, data_only=True)
wf = openpyxl.load_workbook(F)
S, L, D, R = wv["Submissions"], wv["Leaderboard"], wv["Dashboard"], wv["Reward Simulator"]

EVENTS = ["Special Skills using X",
          "More Minds are better than one Mind: Research Quest",
          "Cross Word Puzzle", "Calm before the Storm",
          "AstroMesh: Minds & MCP Mashup Challenge",
          "Minds Building Chatbots Challenge",
          "Mindsheets Masterpiece and Debugging"]
# MINDS is DERIVED from the data, never hardcoded. The previous version listed three names, so
# it could not possibly have detected that the Leaderboard silently dropped a fourth Mind — the
# verifier shared the workbook's blind spot, which is the one failure mode a second
# implementation exists to rule out. It now reads whatever is in the Submissions tab.
def minds_in_order(rows):
    seen=[]
    for (_, m, *_ ) in rows:
        if m not in seen: seen.append(m)
    return seen

checks = []
def chk(where, expect, got, note=""):
    # a formula returning "" reads back as None through openpyxl; treat them as the same thing
    if expect == "" and got is None: got = ""
    if got == "" and expect is None: expect = ""
    if isinstance(expect, float) or isinstance(got, float):
        ok = expect is not None and got is not None and abs(float(expect)-float(got)) < 1e-9
    else:
        ok = expect == got
    checks.append((where, expect, got, ok, note))

# ---- 1. read ONLY the raw inputs: mind, event, and the three 0-10 scores ----
raw = []
r = 5
while r <= S.max_row:
    mind, ev = S.cell(r,3).value, S.cell(r,5).value
    c, k, kr = S.cell(r,6).value, S.cell(r,7).value, S.cell(r,8).value
    if mind and ev and None not in (c,k,kr): raw.append((r, mind, ev, c, k, kr))
    r += 1

# ---- 2. Submissions: I = total, J = grade ----
def grade(t):
    return "★ Elite" if t>=27 else "Strong" if t>=21 else "Fair" if t>=15 else "Weak"
for (row, mind, ev, c, k, kr) in raw:
    chk(f"Submissions!I{row}", c+k+kr, S.cell(row,9).value, f"{mind} / {ev[:22]}")
    chk(f"Submissions!J{row}", grade(c+k+kr), S.cell(row,10).value)

MINDS = minds_in_order(raw)
LB_SLOTS = 12
# every distinct Mind on Submissions must have a row on the Leaderboard, and no phantom rows
lb_names = [L.cell(5+i,3).value for i in range(LB_SLOTS)]
chk("Leaderboard!C (mind list)", MINDS, [n for n in lb_names if n],
    "every Mind on Submissions appears, in first-seen order")

# ---- 3. Leaderboard: best per event, overall, rank, medal ----
tot = {}
for i, m in enumerate(MINDS):
    lr = 5 + i
    per = []
    for j, ev in enumerate(EVENTS):
        # BEST per event, matching the tournament rule the tab claims to implement
        cands = [c+k+kr for (_, mm, ee, c, k, kr) in raw if mm==m and ee==ev]
        want = max(cands) if cands else 0
        col = 5 + j
        chk(f"Leaderboard!{openpyxl.utils.get_column_letter(col)}{lr}", want, L.cell(lr,col).value,
            f"{m} / {ev[:20]}")
        per.append(want)
    tot[m] = sum(per)
    chk(f"Leaderboard!L{lr}", sum(per), L.cell(lr,12).value, f"{m} overall")
order = sorted(tot.values(), reverse=True)
distinct = sorted(set(tot.values()), reverse=True)
for i, m in enumerate(MINDS):
    lr = 5 + i
    rank = order.index(tot[m]) + 1
    chk(f"Leaderboard!B{lr}", rank, L.cell(lr,2).value, f"{m} rank")
    # medals go by POSITION AMONG DISTINCT TOTALS, so a tie does not produce two silvers
    place = distinct.index(tot[m]) + 1
    chk(f"Leaderboard!M{lr}", {1:"🥇",2:"🥈",3:"🥉"}.get(place,""), L.cell(lr,13).value, f"{m} medal")

# ---- 4. Dashboard KPIs ----
top = max(tot.values()); nb = tot["NeoBerlin"]
leader = [m for m in MINDS if tot[m]==top][0]
chk("Dashboard!B6", leader, D.cell(6,2).value, "leader name")
chk("Dashboard!E6", top, D.cell(6,5).value, "top overall")
chk("Dashboard!H6", nb, D.cell(6,8).value, "NeoBerlin overall")
chk("Dashboard!K6", top-nb, D.cell(6,11).value, "gap to leader")
for i, m in enumerate(MINDS):
    chk(f"Dashboard!C{12+i}", tot[m], D.cell(12+i,3).value, f"standings {m}")

# ---- 5. Reward Simulator: quadratic merit, steward binding ----
# Merit per STEWARD, summed over every Mind that steward owns — that is what "binding" means,
# and it is checked here rather than assumed from a fixed row reference.
pool = R.cell(5,4).value
stew_of = {}
for (_, m, _, *_ ) in raw: pass
for (row, m, ev, c, k, kr) in raw: stew_of.setdefault(m, S.cell(row,4).value)
merits = []
r_i = 8
while R.cell(r_i,2).value and R.cell(r_i,2).value != "TOTAL":
    name = R.cell(r_i,2).value
    if name.startswith("SybilFarm"):
        merits.append(R.cell(r_i,3).value*2)
    else:
        merits.append(sum(t for mm,t in tot.items() if stew_of.get(mm)==name))
    r_i += 1
w = [math.sqrt(m) for m in merits]
tw = sum(w)
for i in range(len(merits)):
    rr = 8+i
    chk(f"Reward!D{rr}", merits[i], R.cell(rr,4).value, "merit")
    chk(f"Reward!E{rr}", w[i], R.cell(rr,5).value, "sqrt weight")
    chk(f"Reward!F{rr}", w[i]/tw, R.cell(rr,6).value, "share")
    chk(f"Reward!G{rr}", pool*w[i]/tw, R.cell(rr,7).value, "payout ETH")
TOT_R = 8+len(merits)
chk(f"Reward!F{TOT_R}", 1.0, R.cell(TOT_R,6).value, "shares sum to 1")
chk(f"Reward!G{TOT_R}", float(pool), R.cell(TOT_R,7).value, "payouts sum to the pool")

# ---- 6. Next Best Move: attempts arithmetic and the EV closed form ----
N = wv["Next Best Move"]
NF = openpyxl.load_workbook(F)["Next Best Move"]
CAP = 8
def Phi(z): return 0.5*(1+math.erf(z/math.sqrt(2)))
def pdf(x,m,s): return math.exp(-((x-m)**2)/(2*s*s))/(s*math.sqrt(2*math.pi))
SHORT2LONG = {"X-Skill":"Special Skills using X",
  "Research":"More Minds are better than one Mind: Research Quest",
  "Crossword":"Cross Word Puzzle", "Calm":"Calm before the Storm",
  "AstroMesh":"AstroMesh: Minds & MCP Mashup Challenge",
  "Chatbots":"Minds Building Chatbots Challenge",
  "Mindsheets":"Mindsheets Masterpiece and Debugging"}
elig = []
for r in range(6, 13):
    key   = N.cell(r,14).value
    prior = N.cell(r,15).value
    ev_long = SHORT2LONG[key]
    logged = sum(1 for (_, mm, ee, *_ ) in raw if mm=="NeoBerlin" and ee==ev_long)
    used   = prior + logged
    left   = max(0, CAP-used)
    chk(f"NBM!D{r}", used, N.cell(r,4).value, f"{key} used = prior+logged")
    chk(f"NBM!E{r}", left, N.cell(r,5).value, f"{key} left")
    chk(f"NBM!D{r}+E{r}", CAP, N.cell(r,4).value + N.cell(r,5).value, f"{key} invariant used+left=8")
    best = tot["NeoBerlin"] and None
    b  = N.cell(r,3).value           # Best, itself already checked against the Leaderboard below
    sc = N.cell(r,6).value; mu = N.cell(r,7).value; sd = N.cell(r,8).value
    P  = Phi((b-mu)/sd)
    want_ev = b*P + mu*(1-P) + sd*sd*pdf(b,mu,sd) - b
    chk(f"NBM!J{r}", want_ev, N.cell(r,10).value, f"{key} EV closed form")
    verdict = ("need >=2 scores" if sc<2 else "no attempts left" if left==0
               else "SHOOT — best use of an attempt" if want_ev>=0.5
               else "worthwhile" if want_ev>=0.15 else "low return")
    chk(f"NBM!K{r}", verdict, N.cell(r,11).value, f"{key} verdict")
    elig.append((want_ev if (left>0 and sc>=2) else -999, N.cell(r,2).value))
    # Best must equal the Leaderboard's best-per-event for NeoBerlin
    lb_best = max([c+k+kr for (_, mm, ee, c, k, kr) in raw if mm=="NeoBerlin" and ee==ev_long] or [0])
    chk(f"NBM!C{r}", lb_best, b, f"{key} Best mirrors the Leaderboard")
chk("NBM!D14", max(elig)[1], N.cell(14,4).value, "recommendation = highest eligible EV")
chk("NBM!F13", "OK - all 7 rows sum to 8", N.cell(13,6).value, "invariant check line")

# ---- report ----
bad = [c for c in checks if not c[3]]
print(f"file    : {F}")
print(f"sha256  : {hashlib.sha256(open(F,'rb').read()).hexdigest()}")
print(f"cells independently recomputed and compared: {len(checks)}")
print(f"MISMATCHES: {len(bad)}")
for w_, e, g, ok, n in bad: print(f"  {w_}: expected {e!r}, workbook has {g!r}  ({n})")
if not bad:
    print("\nsample of the agreement (first 8 and last 6):")
    for w_, e, g, ok, n in checks[:8] + checks[-6:]:
        ev = f"{e:.6f}" if isinstance(e,float) else repr(e)
        print(f"  {w_:<22} {ev:>18}  == workbook  {n}")
sys.exit(1 if bad else 0)

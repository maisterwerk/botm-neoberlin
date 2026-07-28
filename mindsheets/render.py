#!/usr/bin/env python3
"""Renders the workbook as compact, readable text: a value grid per tab, then the distinct
formulas with the range they apply to. This is what goes into the submission, so that the whole
artifact can be inspected without downloading a file or following a link."""
import openpyxl, re, hashlib, sys
F="BattleOfTheMinds_Scoreboard.xlsx"
wf=openpyxl.load_workbook(F); wv=openpyxl.load_workbook(F,data_only=True)
def norm(f, r):
    """Replace this row's number inside CELL REFERENCES only, so a formula copied down a column
    collapses to one entry. Matching a bare number would also hit COUNT's letters or a literal
    threshold like 27, so the pattern requires column letters in front of the digits."""
    return re.sub(r'(\$?[A-Z]{1,3}\$?)' + str(r) + r'(?!\d)', r'\1{row}', f)
def fmt(v):
    if v is None: return ""
    if isinstance(v,float):
        return f"{v:.4f}".rstrip("0").rstrip(".") if abs(v)<1e6 else f"{v:.2e}"
    return str(v)
print(f"sha256({F}) = {hashlib.sha256(open(F,'rb').read()).hexdigest()}")
for name in wf.sheetnames:
    a,b = wf[name], wv[name]
    print(f"\n{'='*78}\nTAB  {name}\n{'='*78}")
    # value grid
    blanks=[]
    for r in range(1, a.max_row+1):
        cells=[(a.cell(r,c).coordinate, b.cell(r,c).value) for c in range(1,a.max_column+1)
               if a.cell(r,c).value is not None or b.cell(r,c).value is not None]
        if not cells: continue
        vis=[f"{ref}={fmt(v)}" for ref,v in cells if fmt(v)!=""]
        # An empty data row still carries formulas that evaluate to "" or 0. Printing 26 of
        # them adds nothing, so they are collapsed into one line that says so.
        if vis and not all(re.fullmatch(r"[A-Z]+\d+=0?", x) for x in vis):
            print("  " + "  ".join(vis))
        elif vis:
            blanks.append(r)
    if blanks:
        print(f"  [rows {blanks[0]}-{blanks[-1]}: empty input rows; their formulas return \"\" or 0 "
              f"and are listed once below]")
    # distinct formulas
    seen={}
    for r in range(1,a.max_row+1):
        for c in range(1,a.max_column+1):
            v=a.cell(r,c).value
            if isinstance(v,str) and v.startswith("="):
                key=(openpyxl.utils.get_column_letter(c), norm(v,r))
                seen.setdefault(key,[]).append(r)
    if seen:
        print("  --- formulas ---")
        for (col,f),rows in seen.items():
            rng = f"{col}{rows[0]}" if len(rows)==1 else f"{col}{rows[0]}:{col}{rows[-1]}"
            print(f"  {rng:<12} {f}")
    # ---- the parts the brief explicitly scores: theme, conditional formatting, validation,
    #      number formats and the chart. Values and formulas alone leave half the rubric on
    #      trust, which is the whole thing this rendering exists to avoid.
    if a.sheet_properties.tabColor is not None:
        print(f"  --- tab colour: #{a.sheet_properties.tabColor.rgb} · gridlines "
              f"{'off' if not a.sheet_view.showGridLines else 'on'}"
              f"{' · freeze ' + a.freeze_panes if a.freeze_panes else ''} ---")
    dvs = list(a.data_validations.dataValidation)
    if dvs:
        print("  --- data validation ---")
        for d in dvs:
            rngs = " ".join(str(x) for x in d.sqref.ranges)
            spec = (f"{d.type} {d.operator or ''} {d.formula1 or ''} {d.formula2 or ''}").strip()
            print(f"  {rngs:<14} {spec}   enforced={'yes' if d.showErrorMessage else 'NO (warning only)'}"
                  + (f'  msg="{d.error}"' if d.error else ""))
    cfs = list(a.conditional_formatting)
    if cfs:
        print("  --- conditional formatting ---")
        for cf in cfs:
            for rule in cf.rules:
                bits = [rule.type]
                if rule.formula: bits.append("if " + "; ".join(str(x) for x in rule.formula))
                if rule.type == "colorScale" and rule.colorScale:
                    cs = rule.colorScale
                    bits.append("scale " + " -> ".join("#"+c.rgb[-6:] for c in cs.color))
                if rule.dxf is not None and rule.dxf.fill is not None and rule.dxf.fill.bgColor is not None:
                    bits.append("fill #"+str(rule.dxf.fill.bgColor.rgb)[-6:])
                if rule.dxf is not None and rule.dxf.font is not None and rule.dxf.font.color is not None:
                    bits.append("font #"+str(rule.dxf.font.color.rgb)[-6:])
                print(f"  {str(cf.sqref):<14} {'  '.join(bits)}")
    fmts = {}
    for r_ in range(1, a.max_row+1):
        for c_ in range(1, a.max_column+1):
            nf = a.cell(r_,c_).number_format
            if nf and nf != "General":
                fmts.setdefault(nf, []).append(a.cell(r_,c_).coordinate)
    if fmts:
        print("  --- number formats ---")
        for nf, cs in fmts.items():
            print(f"  {nf:<14} {cs[0]}{'..'+cs[-1] if len(cs)>1 else ''}  ({len(cs)} cells)")
    if getattr(a, "_charts", None):
        for ch in a._charts:
            print(f"  --- chart: {type(ch).__name__} type={getattr(ch,'type',None)} "
                  f"title=\"{ch.title.tx.rich.p[0].r[0].t if ch.title else ''}\" ---")
    hidden = [k for k,v in a.column_dimensions.items() if v.hidden]
    if hidden: print(f"  --- hidden helper columns: {', '.join(hidden)} ---")
